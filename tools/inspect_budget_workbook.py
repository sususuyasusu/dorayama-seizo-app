#!/usr/bin/env python3
"""予算計画原本の構造と表示値を、変更せず確認用JSONへ書き出す。"""
import json
import sys
from pathlib import Path

import openpyxl


def compact(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def main(source, destination):
    formulas = openpyxl.load_workbook(source, data_only=False, read_only=False)
    values = openpyxl.load_workbook(source, data_only=True, read_only=False)
    result = {"source": Path(source).name, "sheets": []}
    for sheet in formulas.worksheets:
        value_sheet = values[sheet.title]
        rows = []
        formula_count = 0
        error_count = 0
        for row_index in range(1, sheet.max_row + 1):
            items = []
            for column_index in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_index, column_index)
                value_cell = value_sheet.cell(row_index, column_index)
                if cell.data_type == "f":
                    formula_count += 1
                rendered = compact(value_cell.value if cell.data_type == "f" else cell.value)
                if isinstance(rendered, str) and rendered.startswith("#"):
                    error_count += 1
                if rendered not in (None, ""):
                    items.append({
                        "cell": cell.coordinate,
                        "value": rendered,
                        "formula": cell.value if cell.data_type == "f" else None,
                    })
            if items:
                rows.append({"row": row_index, "items": items})
        result["sheets"].append({
            "name": sheet.title,
            "maxRow": sheet.max_row,
            "maxColumn": sheet.max_column,
            "merged": [str(value) for value in sheet.merged_cells.ranges],
            "formulaCount": formula_count,
            "errorCount": error_count,
            "rows": rows,
        })
    Path(destination).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
